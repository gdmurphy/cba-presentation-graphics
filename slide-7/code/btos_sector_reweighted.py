#!/usr/bin/env python3
"""
Slide 7: BTOS AI Adoption — Sector-Reweighted for Australia

Reads:
  1. BTOS Sector.xlsx (20251208 vintage) — AI adoption by NAICS sector
  2. SUSB us_state_naics_detailedsizes_2022.xlsx — US receipts by sector
  3. ABS GVA by Industry (hardcoded from Table 5, 2024-25)

Outputs:
  - ai_adoption_sector_reweighted.png — chart with US-weighted and AU-weighted lines

Methodology:
  - US baseline:  r_US = Σ share_US(sector) × r_US(sector)  [SUSB receipts]
  - AU synthetic: r_AU = Σ share_AU(sector) × r_US(sector)  [ABS GVA mapped ANZSIC→NAICS]
  - Only uses sectors present in BOTH BTOS and the weight source

Usage:
  python3 /Users/candaan/Desktop/slide_7/code/btos_sector_reweighted.py
  (or adjust BASE path at top of script if your layout differs)
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
import os

# =========================================================
# Config — UPDATE THESE PATHS to match your machine
# =========================================================
BASE = '/Users/candaan/Desktop/slide_7'

# BTOS Sector data
BTOS_SECTOR_FILE = os.path.join(BASE, 'data', 'Sector.xlsx')

# SUSB receipts data
SUSB_FILE = os.path.join(BASE, 'data', 'us_state_naics_detailedsizes_2022.xlsx')

# Output
OUTPUT_FILE = os.path.join(BASE, 'results', 'ai_adoption_sector_reweighted.png')

# =========================================================
# NAICS sector labels (2-digit codes used in BTOS)
# =========================================================
NAICS_LABELS = {
    11: 'Agriculture',
    21: 'Mining/Oil/Gas',
    22: 'Utilities',
    23: 'Construction',
    31: 'Manufacturing',       # 31-33
    42: 'Wholesale Trade',
    44: 'Retail Trade',        # 44-45
    48: 'Transportation',      # 48-49
    51: 'Information',
    52: 'Finance/Insurance',
    53: 'Real Estate',
    54: 'Professional/Technical',
    55: 'Mgmt of Companies',
    56: 'Administrative/Waste',
    61: 'Education',
    62: 'Health Care',
    71: 'Arts/Entertainment',
    72: 'Accommodation/Food',
    81: 'Other Services',
}

# =========================================================
# ABS GVA by Industry — 2024-25, current prices ($M AUD)
# Source: ABS 5204.0 Table 5, Australian System of National Accounts
# =========================================================
ABS_GVA_2025 = {
    'A': 62986,    # Agriculture, forestry and fishing
    'B': 257665,   # Mining
    'C': 147758,   # Manufacturing
    'D': 59612,    # Electricity, gas, water and waste services
    'E': 198741,   # Construction
    'F': 107677,   # Wholesale trade
    'G': 117706,   # Retail trade
    'H': 67220,    # Accommodation and food services
    'I': 124821,   # Transport, postal and warehousing
    'J': 58727,    # Information media and telecommunications
    'K': 200420,   # Financial and insurance services
    'L': 71072,    # Rental, hiring and real estate services
    'M': 204595,   # Professional, scientific and technical services
    'N': 93072,    # Administrative and support services
    'O': 148089,   # Public administration and safety  [NOT IN BTOS]
    'P': 131642,   # Education and training
    'Q': 229666,   # Health care and social assistance
    'R': 20978,    # Arts and recreation services
    'S': 44897,    # Other services
}

# =========================================================
# ANZSIC division → NAICS 2-digit sector mapping
#
# Both classifications derive from UN ISIC, so divisions
# align well at the top level. Key notes:
#   - ANZSIC D includes waste services; NAICS 22 is Utilities
#     only. Waste is in NAICS 56. We map D→22 since most of
#     its GVA is electricity/gas/water. The mismatch is minor.
#   - ANZSIC O (Public Admin) has no BTOS equivalent — excluded.
#   - NAICS 55 (Mgmt of Companies) has no ANZSIC equivalent —
#     excluded from AU weights; kept in US weights.
# =========================================================
ANZSIC_TO_NAICS = {
    'A': 11,   # Agriculture → Agriculture
    'B': 21,   # Mining → Mining/Oil/Gas
    'C': 31,   # Manufacturing → Manufacturing (31-33)
    'D': 22,   # Electricity/Gas/Water/Waste → Utilities
    'E': 23,   # Construction → Construction
    'F': 42,   # Wholesale Trade → Wholesale Trade
    'G': 44,   # Retail Trade → Retail Trade (44-45)
    'H': 72,   # Accommodation/Food → Accommodation/Food
    'I': 48,   # Transport → Transportation (48-49)
    'J': 51,   # Info Media/Telecom → Information
    'K': 52,   # Financial/Insurance → Finance/Insurance
    'L': 53,   # Rental/RE → Real Estate
    'M': 54,   # Professional/Scientific/Technical → Professional/Technical
    'N': 56,   # Administrative/Support → Administrative/Waste
    # O: Public Admin — NOT IN BTOS, excluded
    'P': 61,   # Education → Education
    'Q': 62,   # Healthcare → Health Care
    'R': 71,   # Arts/Recreation → Arts/Entertainment
    'S': 81,   # Other Services → Other Services
}


# =========================================================
# 1. Compute Australian sector shares (GVA-weighted)
# =========================================================
def compute_au_sector_shares():
    """Map ABS GVA to NAICS sectors and return normalized shares."""
    naics_gva = defaultdict(float)
    for anzsic_div, naics_code in ANZSIC_TO_NAICS.items():
        if anzsic_div in ABS_GVA_2025:
            naics_gva[naics_code] += ABS_GVA_2025[anzsic_div]

    total = sum(naics_gva.values())
    shares = {k: v / total for k, v in naics_gva.items()}
    return shares, naics_gva


# =========================================================
# 2. Compute US sector shares from SUSB receipts
# =========================================================
def compute_us_sector_shares(filepath):
    """Read SUSB and return receipts per 2-digit NAICS sector.

    SUSB NAICS codes can be:
      - '--' = all industries (skip)
      - '11', '21', ... = 2-digit sector
      - '31-33', '44-45', '48-49' = combined range sectors
      - 3+ digit codes = subsectors (skip)
    We want the sector-level total (size_code '01').
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    receipts = defaultdict(float)

    # Map range codes to canonical BTOS 2-digit
    RANGE_MAP = {'31-33': 31, '44-45': 44, '48-49': 48}

    for row in ws.iter_rows(min_row=4):
        state = str(row[0].value).strip() if row[0].value else ''
        naics = str(row[2].value).strip() if row[2].value else ''
        size_raw = str(row[4].value).strip() if row[4].value else ''
        receipt_val = row[11].value

        if state != '00':
            continue

        size_code = size_raw.split(':')[0].strip()
        if size_code != '01':  # total for this sector
            continue

        # Parse NAICS to canonical 2-digit
        if naics in RANGE_MAP:
            naics_int = RANGE_MAP[naics]
        elif naics == '--' or not naics.isdigit():
            continue
        elif len(naics) == 2:
            naics_int = int(naics)
        else:
            continue  # skip 3+ digit subsectors

        r = float(receipt_val) if receipt_val is not None else 0
        receipts[naics_int] += r

    wb.close()

    # Consolidate any remaining split codes (in case stored individually)
    consolidated = defaultdict(float)
    for code, val in receipts.items():
        if code in (32, 33):
            consolidated[31] += val
        elif code == 45:
            consolidated[44] += val
        elif code == 49:
            consolidated[48] += val
        else:
            consolidated[code] += val

    total = sum(consolidated.values())
    if total == 0:
        print("  WARNING: No SUSB sector receipts found! Check file format.")
        return {}, {}

    shares = {k: v / total for k, v in consolidated.items()}
    return shares, consolidated


# =========================================================
# 3. Read BTOS Sector data
# =========================================================
def period_to_date(code):
    """Convert BTOS period code (e.g. 202319) to approximate date."""
    if code is None:
        return None
    code_str = str(int(code))
    if len(code_str) < 5:
        return None
    year = int(code_str[:4])
    fortnight = int(code_str[4:])
    return datetime(year, 1, 1) + timedelta(days=(fortnight - 1) * 14)


def parse_pct(val):
    """Parse a percentage value from BTOS cell. Returns float in [0,1] or None."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace('%', '').strip()
        if val in ('.', '', 'S', 's'):  # S = suppressed
            return None
        try:
            return float(val) / 100
        except ValueError:
            return None
    elif isinstance(val, (int, float)):
        if val > 1:
            return val / 100
        return val
    return None


def read_btos_sector(filepath):
    """Read BTOS Sector.xlsx and return adoption data by sector and period.

    Returns: old_current, new_current, old_future, new_future, dates, period_codes
    Each is {sector_int: {period_idx: rate}}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['Response Estimates']

    header = []
    for row in ws.iter_rows(max_row=1):
        header = [cell.value for cell in row]
        break

    # Columns: Sector, Question ID, Question, Answer ID, Answer, then period codes
    period_codes = header[5:]
    dates = [period_to_date(p) for p in period_codes]

    old_current = defaultdict(dict)
    new_current = defaultdict(dict)
    old_future = defaultdict(dict)
    new_future = defaultdict(dict)

    for row in ws.iter_rows(min_row=2):
        sector_raw = row[0].value
        qid = row[1].value
        qtxt = str(row[2].value) if row[2].value else ''
        aid = row[3].value

        # Only Q7/Q24, Answer=Yes (1)
        if qid not in (7, 24) or aid != 1:
            continue

        # Parse sector code
        try:
            sector = int(sector_raw)
        except (ValueError, TypeError):
            continue

        if sector not in NAICS_LABELS:
            continue

        is_old = 'producing goods or services' in qtxt
        if qid == 7:
            target = old_current if is_old else new_current
        else:
            target = old_future if is_old else new_future

        for i, cell in enumerate(list(row)[5:]):
            val = parse_pct(cell.value)
            if val is not None:
                target[sector][i] = val

    wb.close()
    return old_current, new_current, old_future, new_future, dates, period_codes


# =========================================================
# 4. Compute weighted adoption lines (sector-level)
# =========================================================
def compute_sector_weighted_line(data, weights, min_coverage=0.70):
    """
    Given {sector: {period_idx: rate}} and {sector: weight},
    compute weighted-average rate per period.

    Unlike empsize weighting (where we require ALL buckets), sector data
    has more suppression ('S' values). We include a period if sectors
    covering >= min_coverage of total weight have data.
    The weights are renormalized to sum to 1 over available sectors.
    """
    all_periods = set()
    for sector_data in data.values():
        all_periods.update(sector_data.keys())

    result = {}
    for p in sorted(all_periods):
        num = 0
        den = 0
        for sector, weight in weights.items():
            if sector in data and p in data[sector]:
                num += data[sector][p] * weight
                den += weight

        if den >= min_coverage * sum(weights.values()) and den > 0:
            result[p] = num / den

    return result


# =========================================================
# 5. Main
# =========================================================
def main():
    print("=" * 65)
    print("BTOS AI Adoption — Sector-Reweighted (Australia vs US)")
    print("=" * 65)

    # --- Australian GVA shares ---
    print("\n[1/4] Computing Australian sector shares (ABS GVA 2024-25)...")
    au_shares, au_gva = compute_au_sector_shares()

    # --- US sector shares ---
    print("[2/4] Reading US sector shares (SUSB 2022 receipts)...")
    us_shares, us_receipts = compute_us_sector_shares(SUSB_FILE)

    # Print comparison
    print(f"\n  {'NAICS':<6} {'Sector':<24} {'AU Share':>9} {'US Share':>9} {'Diff':>7}")
    print(f"  {'-'*6} {'-'*24} {'-'*9} {'-'*9} {'-'*7}")
    for naics in sorted(set(list(au_shares.keys()) + list(us_shares.keys()))):
        name = NAICS_LABELS.get(naics, '???')
        au = au_shares.get(naics, 0)
        us = us_shares.get(naics, 0)
        diff = au - us
        flag = ' <--' if abs(diff) > 0.03 else ''
        print(f"  {naics:<6} {name:<24} {au:>8.1%} {us:>8.1%} {diff:>+6.1%}{flag}")

    # --- BTOS sector data ---
    print(f"\n[3/4] Reading BTOS sector data...")
    old_current, new_current, old_future, new_future, dates, period_codes = \
        read_btos_sector(BTOS_SECTOR_FILE)

    for label, data in [("Old Q7 (current, 'production')", old_current),
                        ("New Q7 (current, 'any function')", new_current),
                        ("Old Q24 (future, 'production')", old_future),
                        ("New Q24 (future, 'any function')", new_future)]:
        n_sectors = len(data)
        n_periods = max((len(v) for v in data.values()), default=0)
        print(f"  {label}: {n_sectors} sectors, ~{n_periods} periods")

    # --- Compute weighted lines ---
    print(f"\n[4/4] Computing weighted lines...")

    # US sector-receipt weighted (US baseline)
    old_curr_us = compute_sector_weighted_line(old_current, us_shares)
    new_curr_us = compute_sector_weighted_line(new_current, us_shares)
    old_fut_us = compute_sector_weighted_line(old_future, us_shares)
    new_fut_us = compute_sector_weighted_line(new_future, us_shares)

    # AU GVA-sector weighted (synthetic Australian benchmark)
    old_curr_au = compute_sector_weighted_line(old_current, au_shares)
    new_curr_au = compute_sector_weighted_line(new_current, au_shares)
    old_fut_au = compute_sector_weighted_line(old_future, au_shares)
    new_fut_au = compute_sector_weighted_line(new_future, au_shares)

    print(f"  US old current: {len(old_curr_us)} periods")
    print(f"  AU old current: {len(old_curr_au)} periods")
    print(f"  US new current: {len(new_curr_us)} periods")
    print(f"  AU new current: {len(new_curr_au)} periods")

    # Print some sample values
    if old_curr_us and old_curr_au:
        last_p = max(old_curr_us.keys())
        d = dates[last_p]
        print(f"\n  Latest old Q7 period ({d.strftime('%b %Y') if d else '?'}):")
        print(f"    US sector-weighted: {old_curr_us[last_p]*100:.1f}%")
        print(f"    AU sector-weighted: {old_curr_au[last_p]*100:.1f}%")

    # =========================================================
    # Plot
    # =========================================================
    def get_xy(line_data):
        """Convert {period_idx: rate} to sorted (dates, values%) arrays."""
        periods = sorted(line_data.keys())
        d = [dates[p] for p in periods if dates[p] is not None]
        v = [line_data[p] * 100 for p in periods if dates[p] is not None]
        return d, v

    fig, ax = plt.subplots(figsize=(14, 7))

    # --- Current use: US sector-weighted ---
    d, v = get_xy(old_curr_us)
    ax.plot(d, v, color='#1f4e79', linewidth=2, solid_capstyle='round',
            label='AI Current Use (US sector mix)')
    d, v = get_xy(new_curr_us)
    if d:
        ax.plot(d, v, color='#1f4e79', linewidth=2, solid_capstyle='round')

    # --- Current use: AU sector-weighted ---
    d, v = get_xy(old_curr_au)
    ax.plot(d, v, color='#27ae60', linewidth=2.5, solid_capstyle='round',
            label='AI Current Use (AU sector mix)')
    d, v = get_xy(new_curr_au)
    if d:
        ax.plot(d, v, color='#27ae60', linewidth=2.5, solid_capstyle='round')

    # --- Future use: US sector-weighted ---
    d, v = get_xy(old_fut_us)
    ax.plot(d, v, color='#1f4e79', linewidth=1.5, linestyle='--', alpha=0.7,
            label='AI Use Next 6 Months (US sector mix)')
    d, v = get_xy(new_fut_us)
    if d:
        ax.plot(d, v, color='#1f4e79', linewidth=1.5, linestyle='--', alpha=0.7)

    # --- Future use: AU sector-weighted ---
    d, v = get_xy(old_fut_au)
    ax.plot(d, v, color='#27ae60', linewidth=1.5, linestyle='--', alpha=0.7,
            label='AI Use Next 6 Months (AU sector mix)')
    d, v = get_xy(new_fut_au)
    if d:
        ax.plot(d, v, color='#27ae60', linewidth=1.5, linestyle='--', alpha=0.7)

    # --- Data break annotation (govt shutdown: Oct 6 - Nov 16, 2025) ---
    break_date = datetime(2025, 10, 15)
    ax.annotate('Data\nBreak', xy=(break_date, 2), fontsize=9, ha='center',
                fontweight='bold', color='#333333',
                arrowprops=dict(arrowstyle='->', color='#333333', lw=1.5),
                xytext=(break_date, 5))

    # --- Formatting ---
    ax.set_title('Artificial Intelligence Adoption', fontsize=16, fontweight='bold', pad=15)
    ax.set_ylabel('AI Use', fontsize=12)
    ax.set_ylim(0, 30)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y:.0f}%'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
    ax.legend(loc='upper left', fontsize=9, framealpha=0.9)
    ax.grid(True, alpha=0.2)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # --- Footnote ---
    fig.text(0.5, -0.02,
             'Note: US sector mix weighted by SUSB 2022 receipts; AU sector mix weighted by ABS GVA 2024-25.\n'
             'ANZSIC→NAICS mapping at division level. BTOS adoption rates are US firms only.\n'
             'Source: Census Bureau BTOS, SUSB 2022, ABS 5204.0 Table 5.',
             ha='center', fontsize=8, color='gray', style='italic')

    plt.tight_layout()
    plt.savefig(OUTPUT_FILE, dpi=200, bbox_inches='tight')
    print(f"\nSaved chart to: {OUTPUT_FILE}")
    plt.show()


if __name__ == '__main__':
    main()