#!/usr/bin/env python3
"""
Slide 7: BTOS AI Adoption — Expenditure-Reweighted Benchmark

Reads:
  1. BTOS Employment Size Class.xlsx (20251208 vintage) — AI adoption by firm size
  2. SUSB us_state_naics_detailedsizes_2022.xlsx — receipts & firm counts by size

Outputs:
  - ai_adoption_reweighted.png — chart with original and reweighted lines

Usage:
  cd /Users/candaan/Desktop/slide_7/data
  python3 btos_reweighted.py
"""

import openpyxl
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
import os

# =========================================================
# Config
# =========================================================
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
BTOS_FILE = os.path.join(DATA_DIR, 'Employment Size Class.xlsx')
SUSB_FILE = os.path.join(DATA_DIR, 'us_state_naics_detailedsizes_2022.xlsx')
OUTPUT_FILE = os.path.join(DATA_DIR, 'ai_adoption_reweighted.png')

EMPSIZE_LABELS = {
    'A': '1-4',
    'B': '5-9',
    'C': '10-19',
    'D': '20-49',
    'E': '50-99',
    'F': '100-249',
    'G': '250+',
}

# Mapping from SUSB size-code prefix (e.g. '02') to BTOS empsize letter.
# Code '16' (200-299) straddles the 250 boundary: split 50/50 between F and G.
SUSB_TO_BTOS = {
    '02': 'A',   # <5
    '03': 'B',   # 5-9
    '04': 'C',   # 10-14  (part of C = 10-19)
    '05': 'C',   # 15-19
    '07': 'D',   # 20-24  (parts of D = 20-49)
    '08': 'D',   # 25-29
    '09': 'D',   # 30-34
    '10': 'D',   # 35-39
    '11': 'D',   # 40-49
    '12': 'E',   # 50-74  (parts of E = 50-99)
    '13': 'E',   # 75-99
    '14': 'F',   # 100-149 (parts of F = 100-249)
    '15': 'F',   # 150-199
    # '16' handled specially (200-299 split)
    '17': 'G',   # 300-399 (parts of G = 250+)
    '18': 'G',   # 400-499
    '20': 'G',   # 500-749
    '21': 'G',   # 750-999
    '22': 'G',   # 1000-1499
    '23': 'G',   # 1500-1999
    '24': 'G',   # 2000-2499
    '25': 'G',   # 2500-4999
    '26': 'G',   # 5000+
}


# =========================================================
# 1. Read SUSB — receipts and firm counts by BTOS bucket
# =========================================================
def read_susb(filepath):
    """Read SUSB and return receipts and firm counts per BTOS empsize bucket."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    receipts = defaultdict(float)
    firms = defaultdict(float)

    for row in ws.iter_rows(min_row=4):
        state = str(row[0].value).strip() if row[0].value else ''
        naics = str(row[2].value).strip() if row[2].value else ''
        size_raw = str(row[4].value).strip() if row[4].value else ''
        firm_count = row[5].value
        receipt_val = row[11].value

        # Filter to US total, all industries
        if state != '00' or naics != '--':
            continue

        # Extract size code (e.g. '02' from '02: <5 employees')
        size_code = size_raw.split(':')[0].strip()

        # Skip totals and subtotals
        if size_code in ('01', '06', '19'):
            continue

        # Get numeric values (handle None)
        r = float(receipt_val) if receipt_val is not None else 0
        f = float(firm_count) if firm_count is not None else 0

        if size_code == '16':
            # 200-299 straddles 250: split 50/50 between F and G
            receipts['F'] += r / 2
            receipts['G'] += r / 2
            firms['F'] += f / 2
            firms['G'] += f / 2
        elif size_code in SUSB_TO_BTOS:
            btos_key = SUSB_TO_BTOS[size_code]
            receipts[btos_key] += r
            firms[btos_key] += f

    wb.close()

    # Convert to shares
    total_r = sum(receipts.values())
    total_f = sum(firms.values())
    receipt_shares = {k: v / total_r for k, v in receipts.items()}
    firm_shares = {k: v / total_f for k, v in firms.items()}

    return receipt_shares, firm_shares, receipts, firms


# =========================================================
# 2. Read BTOS — AI adoption by empsize and period
# =========================================================
def period_to_date(code):
    """Convert BTOS period code (e.g. 202319) to approximate date."""
    if code is None:
        return None
    code_str = str(int(code))
    if len(code_str) < 5:
        return None
    year = int(code_str[:4])
    fortnight = int(code_str[4:])
    # Each fortnight ~ 14 days from start of year
    return datetime(year, 1, 1) + timedelta(days=(fortnight - 1) * 14)


def parse_pct(val):
    """Parse a percentage value from BTOS cell. Returns float in [0,1] or None."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace('%', '').strip()
        if val == '.' or val == '':
            return None
        return float(val) / 100
    elif isinstance(val, (int, float)):
        if val > 1:  # Stored as e.g. 5.4 meaning 5.4%
            return val / 100
        return val
    return None


def read_btos(filepath):
    """Read BTOS and return adoption data for Q7 and Q24, old and new wordings."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['Response Estimates']

    # Get header — first 5 cols are metadata, rest are period codes
    header = []
    for row in ws.iter_rows(max_row=1):
        header = [cell.value for cell in row]
        break

    period_codes = header[5:]
    dates = [period_to_date(p) for p in period_codes]

    # Data structures: {empsize: {period_idx: rate}}
    old_current = defaultdict(dict)   # Q7 old "producing goods or services"
    new_current = defaultdict(dict)   # Q7 new "any business functions"
    old_future = defaultdict(dict)    # Q24 old wording
    new_future = defaultdict(dict)    # Q24 new wording

    for row in ws.iter_rows(min_row=2):
        empsize = row[0].value
        qid = row[1].value
        qtxt = str(row[2].value) if row[2].value else ''
        aid = row[3].value

        # Only Q7/Q24, Answer=Yes (1)
        if qid not in (7, 24) or aid != 1:
            continue
        if empsize not in EMPSIZE_LABELS:
            continue

        is_old = 'producing goods or services' in qtxt
        if qid == 7:
            target = old_current if is_old else new_current
        else:
            target = old_future if is_old else new_future

        for i, cell in enumerate(list(row)[5:]):
            val = parse_pct(cell.value)
            if val is not None:
                target[empsize][i] = val

    wb.close()
    return old_current, new_current, old_future, new_future, dates, period_codes


# =========================================================
# 3. Compute weighted adoption lines
# =========================================================
def compute_weighted_line(data, weights):
    """
    Given {empsize: {period_idx: rate}} and {empsize: weight},
    compute weighted-average rate per period.
    Only includes a period if ALL 7 size classes have data for it.
    """
    all_periods = set()
    for emp_data in data.values():
        all_periods.update(emp_data.keys())

    result = {}
    for p in sorted(all_periods):
        num = 0
        den = 0
        all_present = True
        for emp in EMPSIZE_LABELS:
            if emp in data and p in data[emp] and emp in weights:
                num += data[emp][p] * weights[emp]
                den += weights[emp]
            else:
                all_present = False
                break
        if all_present and den > 0:
            result[p] = num / den

    return result


# =========================================================
# 4. Main
# =========================================================
def main():
    print("=" * 60)
    print("BTOS AI Adoption — Expenditure-Reweighted Benchmark")
    print("=" * 60)

    # --- SUSB ---
    print("\n[1/4] Reading SUSB data...")
    receipt_shares, firm_shares, receipts_raw, firms_raw = read_susb(SUSB_FILE)

    print("\n  SUSB Weights Summary:")
    print(f"  {'Bucket':<12} {'Receipt Share':>14} {'Firm Share':>12}")
    print(f"  {'-'*12} {'-'*14} {'-'*12}")
    for emp in sorted(EMPSIZE_LABELS.keys()):
        print(f"  {emp} ({EMPSIZE_LABELS[emp]:>5s})   "
              f"{receipt_shares.get(emp, 0):>13.1%}  "
              f"{firm_shares.get(emp, 0):>11.1%}")

    # --- BTOS ---
    print("\n[2/4] Reading BTOS data...")
    old_current, new_current, old_future, new_future, dates, period_codes = read_btos(BTOS_FILE)

    for label, data in [("Old Q7 (current, 'production')", old_current),
                        ("New Q7 (current, 'any function')", new_current),
                        ("Old Q24 (future, 'production')", old_future),
                        ("New Q24 (future, 'any function')", new_future)]:
        n_sizes = len(data)
        n_periods = max((len(v) for v in data.values()), default=0)
        print(f"  {label}: {n_sizes} size classes, ~{n_periods} periods")

    # --- Compute weighted lines ---
    print("\n[3/4] Computing weighted lines...")

    # Firm-share weighted ~ original BTOS national line
    old_curr_firm = compute_weighted_line(old_current, firm_shares)
    new_curr_firm = compute_weighted_line(new_current, firm_shares)
    old_fut_firm = compute_weighted_line(old_future, firm_shares)
    new_fut_firm = compute_weighted_line(new_future, firm_shares)

    # Receipt-share weighted = our reweighted benchmark
    old_curr_rcpt = compute_weighted_line(old_current, receipt_shares)
    new_curr_rcpt = compute_weighted_line(new_current, receipt_shares)
    old_fut_rcpt = compute_weighted_line(old_future, receipt_shares)
    new_fut_rcpt = compute_weighted_line(new_future, receipt_shares)

    print(f"  Old current use: {len(old_curr_firm)} periods")
    print(f"  New current use: {len(new_curr_firm)} periods")
    print(f"  Old future use:  {len(old_fut_firm)} periods")
    print(f"  New future use:  {len(new_fut_firm)} periods")

    # --- Plot ---
    print("\n[4/4] Plotting...")

    def get_xy(line_data):
        """Convert {period_idx: rate} to sorted (dates, values%) arrays."""
        periods = sorted(line_data.keys())
        d = [dates[p] for p in periods if dates[p] is not None]
        v = [line_data[p] * 100 for p in periods if dates[p] is not None]
        return d, v

    fig, ax = plt.subplots(figsize=(14, 7))

    # --- Current use: firm-weighted (original proxy) ---
    d, v = get_xy(old_curr_firm)
    ax.plot(d, v, color='#1f4e79', linewidth=2, solid_capstyle='round',
            label='AI Current Use (firm-weighted)')
    d, v = get_xy(new_curr_firm)
    if d:
        ax.plot(d, v, color='#1f4e79', linewidth=2, solid_capstyle='round')

    # --- Current use: receipt-weighted (our benchmark) ---
    d, v = get_xy(old_curr_rcpt)
    ax.plot(d, v, color='#c0392b', linewidth=2, solid_capstyle='round',
            label='AI Current Use (receipt-weighted)')
    d, v = get_xy(new_curr_rcpt)
    if d:
        ax.plot(d, v, color='#c0392b', linewidth=2, solid_capstyle='round')

    # --- Future use: firm-weighted ---
    d, v = get_xy(old_fut_firm)
    ax.plot(d, v, color='#1f4e79', linewidth=1.5, linestyle='--', alpha=0.7,
            label='AI Use Next 6 Months (firm-weighted)')
    d, v = get_xy(new_fut_firm)
    if d:
        ax.plot(d, v, color='#1f4e79', linewidth=1.5, linestyle='--', alpha=0.7)

    # --- Future use: receipt-weighted ---
    d, v = get_xy(old_fut_rcpt)
    ax.plot(d, v, color='#c0392b', linewidth=1.5, linestyle='--', alpha=0.7,
            label='AI Use Next 6 Months (receipt-weighted)')
    d, v = get_xy(new_fut_rcpt)
    if d:
        ax.plot(d, v, color='#c0392b', linewidth=1.5, linestyle='--', alpha=0.7)

    # --- Data break annotation (govt shutdown: Oct 6 - Nov 16, 2025) ---
    break_date = datetime(2025, 10, 15)
    ylim_current = ax.get_ylim()
    ax.annotate('Data\nBreak', xy=(break_date, 2), fontsize=9, ha='center',
                fontweight='bold', color='#333333',
                arrowprops=dict(arrowstyle='->', color='#333333', lw=1.5),
                xytext=(break_date, 5))

    # --- Formatting ---
    ax.set_title('Artificial Intelligence Adoption', fontsize=16, fontweight='bold', pad=15)
    ax.set_ylabel('AI Use', fontsize=12)
    ax.set_ylim(0, 30)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f'{y:.0f}%'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
    ax.legend(loc='upper left', fontsize=9, framealpha=0.9)
    ax.grid(True, alpha=0.2)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # --- Footnote ---
    fig.text(0.5, -0.02,
             'Note: Firm-weighted uses SUSB 2022 firm counts; receipt-weighted uses SUSB 2022 total receipts.\n'
             'SUSB 200-299 employee bucket split 50/50 at 250 boundary. '
             'Source: Census Bureau BTOS, SUSB 2022.',
             ha='center', fontsize=8, color='gray', style='italic')

    plt.tight_layout()
    plt.savefig(OUTPUT_FILE, dpi=200, bbox_inches='tight')
    print(f"\nSaved chart to: {OUTPUT_FILE}")
    plt.show()


if __name__ == '__main__':
    main()
